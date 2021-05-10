import openpyxl
from PIL import Image
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
import sys

args = sys.argv
path_file = args[1]
file_sheet = args[2]
file_name = args[3]
date = args[4]

pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))

#import excel file
wb = openpyxl.load_workbook(path_file, data_only=True)
sheet = wb[file_sheet]

#Page Size
pg_width ,pg_height= A4 

#Specify logo size
im = Image.open("./img/Leo'sLogo.jpeg")
width, height = im.size
ratio = width/height
image_width = 100
image_height = image_width/ratio

#Global variable
margin = 10
paragraph = 4

def getLcolIndex():
    i = 1
    while (sheet.cell(row = 3, column = i ).value != None):
        i+= 1
    return i

#fcol - the index where the items start
#lcol - the last index where the items end
def getTotalIndex():
    goal = "total"
    i = 1
    while i < 30: # Keep this cuz not I do not want to get an infinite loop
        # print(i)
        if (sheet.cell(row = 1,column = i).value == None):
            i += 1
        elif (sheet.cell(row = 1,column = i).value.lower() == goal):
            return i
        else:
            i+=1
    return i

def getLastIndex():
    i = 1
    while (sheet.cell(row = i, column = 1 ).value != None):
        i+= 1
    return i

end_index = getLastIndex() - 1
total = getTotalIndex()

fcol = 5 #fcol is always 5
lcol = getLcolIndex() - 1

if (len(args) != 5):
    print('Error')
    exit()



#invoice related information
company_name = "Leo's cuisine"

def items(row, fcol, lcol, tab):
    output = []
    innerlst = []
    for i in range(fcol, lcol + 2):
      
        if (tab.cell(row = 1, column = i).value != None and innerlst != []):
            output.append(innerlst)
            innerlst = []
        if (innerlst == []):
            innerlst.append(tab.cell(row = 1, column = i).value)
        
        innerlst.append(tab.cell(row = row, column = i).value)
    output.append(innerlst)
 
    return output
def atleastone(lst1, lst2):
    # for i in lst1:
    #     if i in lst2:
    for i in lst2:
        for j in lst1:
            if j in i:
                return True
    return False
def create_invoice():
    c = canvas.Canvas(file_name + '.pdf')
    extra_dic = {}
    extra_dic['fried'] = ['gato arouille', 'fried noodles', 'catless']
    extra_dic['boulette'] = ['kniouk yens', 'sao mai', 'long']
    print(sheet)
    print("TOtal index " + str(getTotalIndex()) + "ad")
    print("\nLast index" + str(getLastIndex()))
    print("\nLast col index" + str(getLcolIndex()))

    for i in range(3, int(end_index) + 1):
        #Accumulator
        products = []
        #Reading values from excel file
        expected_Time = sheet.cell(row = i, column = 2).value
        Client_name = sheet.cell(row = i, column = 3).value
        address = sheet.cell(row=i, column= 4).value
        
        sales = items(i, int(fcol),int(lcol), sheet)
        Total = sheet.cell(row = i, column = int(total)).value
        print(sheet.cell(row = 1, column = int(total)).value)
        #Creating the pdf file
        
        
        c.setPageSize((pg_width, pg_height))

        #Draw logo
        y = pg_height - image_height - margin
        c.drawInlineImage("./img/Leo'sLogo.jpeg", (pg_width - image_width)/2, y, image_width, image_height)
        #Invoice
        c.setFont('Arial', 40)
        text = 'INVOICE'
        text_width = stringWidth(text, 'Arial', 40)
        y -= margin * 4
        c.drawString((pg_width - text_width)/2, y,text)

        x = 8 * margin
        y -= margin * paragraph

        #Company Name
        c.setFont('Arial', 20)
        c.drawString(x,y,'Date: ' + date)
        y -= margin * paragraph

        #Pickup Time
        c.setFont('Arial', 20)
        c.drawString(x,y,'Pick Up Time: ' + str(expected_Time))
        y -= margin * paragraph


        #Customer 
        c.drawString(x,y, 'Name: ' + Client_name)
        y -= margin * paragraph

        #Location 
        c.drawString(x,y, 'Location: ' + address )
        y -= margin * paragraph + 8

        #Sales items
        c.setFont('Arial', 30)
        quantity = 'Qty'
        name = 'Product'
        unitprice = 'Price'
        Totalstr = 'Amt'
        quantity_width = stringWidth(quantity, 'Arial', 30)
        name_width = stringWidth(name, 'Arial', 30)
        unitprice_width = stringWidth(unitprice, 'Arial', 30)
        Total_width = stringWidth(Totalstr, 'Arial', 30)
        
        xsection = pg_width - x*2
        divsection = (xsection - Total_width)/4
        section = divsection

        q_section = section
        c.drawString(section, y,quantity)
        section += divsection

        n_section = section
        c.drawString(section, y,name)
        section += divsection + name_width - Total_width + margin*2

        u_section = section
        c.drawString(section, y,unitprice)
        section += divsection

        Total_section = section
        c.drawString(section, y,Totalstr)

        y -= margin * paragraph 
        print(sales)
        for i in sales:
            if (i[1] != 0 and i[2] != 0 and i[1] != None and i[2] != None):
                c.setFont('Arial', 20)
                # print('l'+i[0].lower().strip() + 'f' + extra_dic['boulette'][0] +'l', extra_dic['boulette'], i[0].lower())
                c.drawString(q_section,y,str(i[1]))
                if ('\n' in i[0]):
                    nlindex = i[0].index('\n')
                    c.drawString(n_section,y,i[0][:nlindex])
                    y -= margin * paragraph 
                    c.drawString(n_section,y,i[0][nlindex + 1 :])
                    y += margin * paragraph 
                else:
                    c.drawString(n_section,y,i[0])
                products.append(i[0])
                c.drawString(u_section,y,str(i[2]/i[1]))
                c.drawString(Total_section,y,str(i[2]))
                y -= margin * paragraph
                if ('\n' in i[0]):
                    y -= margin * paragraph


        #Free Extras
        c.setFont('Arial', 25)
        c.drawString(x,y,"Free extras:")
        y -= margin * paragraph

        c.setFont('Arial', 20)
        c.drawString(x,y,"Chilli paste")
        y -= margin * paragraph

        if (atleastone(extra_dic['boulette'], [x.lower().strip() for x in products])):
            c.setFont('Arial', 20)
            c.drawString(x,y,"Soup")
            y -= margin * paragraph

            c.setFont('Arial', 20)
            c.drawString(x,y,"Special Sauce")
            y -= margin * paragraph
        if (atleastone(extra_dic['fried'], [x.lower().strip() for x in products])):
            c.setFont('Arial', 20)
            c.drawString(x,y,"Garlic Sauce")
            y -= margin * paragraph





        #Total price
        c.setFont('Arial', 30)
        c.drawString(x + margin*25,49, 'Total: ')
        c.setFont('Arial', 25)
        c.drawString(x + margin*25 + stringWidth('Total: ', 'Arial', 30), 50,  str(Total))

        y -= margin * paragraph

        c.setFont('Arial', 10)
        notice = 'A fee may apply if cancellation notice is given less than 3 days'
        notice_width = stringWidth(notice, 'Arial', 10)
        c.drawString((pg_width - notice_width)/2,10, notice)
        y -= margin * paragraph

        



        c.showPage()

    c.save()

create_invoice()
